import openpyxl
import pandas as pd
import base64
import io
from PIL import Image as PILImage
from pathlib import Path

def extract_images_from_reference_sheet(file_path, sheet_name="Справочник пробирок"):
    """
    Извлекает изображения из листа 'Справочник пробирок' и создает словарь 
    соответствий между типами контейнеров и base64 изображениями
    """
    # Загружаем рабочую книгу
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    # Словарь для хранения соответствий
    image_dict = {}
    
    # Читаем типы контейнеров из столбца A (начиная со строки 2)
    container_types = {}
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet[f'A{row}'].value
        if cell_value and cell_value.strip():
            container_types[row] = cell_value.strip()
    
    print(f"Найдено типов контейнеров: {len(container_types)}")
    for row, container_type in container_types.items():
        print(f"  Строка {row}: {container_type}")
    
    # Получаем все изображения из листа
    images = worksheet._images
    print(f"Найдено изображений в листе: {len(images)}")
    
    # Обрабатываем каждое изображение
    for img_idx, img in enumerate(images):
        try:
            # Получаем позицию изображения
            if hasattr(img.anchor, '_from'):
                # Для newer versions of openpyxl
                row_idx = img.anchor._from.row + 1
                col_idx = img.anchor._from.col + 1
            else:
                # Для older versions
                row_idx = img.anchor.row if hasattr(img.anchor, 'row') else 1
                col_idx = img.anchor.col if hasattr(img.anchor, 'col') else 1
            
            print(f"Изображение {img_idx + 1}: позиция строка {row_idx}, столбец {col_idx}")
            
            # Находим ближайший тип контейнера
            container_type = None
            min_distance = float('inf')
            
            for row_num, c_type in container_types.items():
                distance = abs(row_num - row_idx)
                if distance < min_distance:
                    min_distance = distance
                    container_type = c_type
            
            if container_type and min_distance <= 2:  # Допускаем погрешность в 2 строки
                try:
                    # Получаем данные изображения
                    # img.ref может быть BytesIO объектом или байтами
                    if hasattr(img.ref, 'read'):
                        # Если это BytesIO объект
                        img.ref.seek(0)  # Переходим к началу
                        img_data = img.ref.read()
                    else:
                        # Если это уже байты
                        img_data = img.ref
                    
                    # Конвертируем в PIL Image и затем в base64
                    pil_image = PILImage.open(io.BytesIO(img_data))
                    
                    # Изменяем размер изображения для оптимизации (опционально)
                    max_size = (300, 300)
                    pil_image.thumbnail(max_size, PILImage.Resampling.LANCZOS)
                    
                    # Конвертируем в base64
                    buffered = io.BytesIO()
                    pil_image.save(buffered, format="PNG")
                    img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                    
                    image_dict[container_type] = img_base64
                    print(f"  Сопоставлено: {container_type} -> изображение {len(img_base64)} символов")
                    
                except Exception as e:
                    print(f"  Ошибка при обработке изображения для {container_type}: {e}")
                    # Попробуем альтернативный способ извлечения
                    try:
                        # Альтернативный способ - через _data атрибут
                        if hasattr(img, '_data'):
                            img_data = img._data
                        elif hasattr(img, 'path'):
                            # Если изображение хранится как файл
                            with open(img.path, 'rb') as f:
                                img_data = f.read()
                        else:
                            print(f"  Не удалось найти данные изображения для {container_type}")
                            continue
                            
                        pil_image = PILImage.open(io.BytesIO(img_data))
                        max_size = (300, 300)
                        pil_image.thumbnail(max_size, PILImage.Resampling.LANCZOS)
                        
                        buffered = io.BytesIO()
                        pil_image.save(buffered, format="PNG")
                        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                        
                        image_dict[container_type] = img_base64
                        print(f"  Сопоставлено (альтернативный способ): {container_type} -> изображение {len(img_base64)} символов")
                        
                    except Exception as e2:
                        print(f"  Альтернативный способ тоже не сработал для {container_type}: {e2}")
                        continue
            else:
                print(f"  Не найден подходящий тип контейнера для изображения на позиции {row_idx}")
                
        except Exception as e:
            print(f"Ошибка при обработке изображения {img_idx + 1}: {e}")
            continue
    
    workbook.close()
    return image_dict

def find_matching_image(container_type, image_dict):
    """
    Находит подходящее изображение по вхождению названия
    """
    if not container_type or pd.isna(container_type):
        return None
        
    container_type_clean = str(container_type).strip().lower()
    
    # Сначала ищем точное совпадение
    for ref_type, base64_img in image_dict.items():
        if container_type_clean == ref_type.lower().strip():
            return base64_img
    
    # Затем ищем по вхождению - название из data входит в название из справочника
    for ref_type, base64_img in image_dict.items():
        ref_type_clean = ref_type.lower().strip()
        if container_type_clean in ref_type_clean:
            return base64_img
    
    # Затем ищем обратное вхождение - название из справочника входит в название из data
    for ref_type, base64_img in image_dict.items():
        ref_type_clean = ref_type.lower().strip()
        if ref_type_clean in container_type_clean:
            return base64_img
    
    # Ищем по ключевым словам
    keywords_mapping = {
        'белая': ['белая', 'белой'],
        'красная': ['красная', 'красной', 'гель', 'гелем'],
        'сиреневая': ['сиреневая', 'сиреневой', 'розовая', 'розовой', 'эдта'],
        'желтая': ['желтая', 'желтой', 'моча', 'мочи', 'консервант'],
        'стерильный': ['стерильный', 'стерильной', 'спирт'],
        'кал': ['кал', 'кала', 'ложечка', 'ложечкой'],
        'микро': ['микро', 'транспорт'],
        'стекло': ['стекло', 'предметное'],
        'флакон': ['флакон', 'юнона', 'детский'],
        'amies': ['amies', 'оранжевая', 'оранжевой'],
        'гистолог': ['гистолог', 'histopot'],
        'парафин': ['парафин', 'блок']
    }
    
    for ref_type, base64_img in image_dict.items():
        ref_type_clean = ref_type.lower().strip()
        for keyword, synonyms in keywords_mapping.items():
            if any(syn in container_type_clean for syn in synonyms) and keyword in ref_type_clean:
                return base64_img
    
    return None

def add_base64_to_data_sheet(file_path, image_dict, data_sheet="data"):
    """
    Добавляет столбец с base64 изображениями в лист 'data'
    """
    # Читаем данные из листа 'data'
    df = pd.read_excel(file_path, sheet_name=data_sheet)
    
    print(f"Загружено строк данных: {len(df)}")
    print(f"Столбцы: {list(df.columns)}")
    
    # Создаем новый столбец с base64 изображениями используя функцию поиска по вхождению
    df['container_image_base64'] = df['container_type'].apply(lambda x: find_matching_image(x, image_dict))
    
    # Проверяем результат
    matched = df['container_image_base64'].notna().sum()
    total = len(df)
    print(f"Сопоставлено изображений: {matched} из {total}")
    
    # Показываем детальную статистику сопоставления
    print("\nДетальная статистика сопоставления:")
    container_counts = df['container_type'].value_counts()
    for container_type, count in container_counts.items():
        matched_image = find_matching_image(container_type, image_dict)
        if matched_image:
            # Находим какое именно изображение было сопоставлено
            matched_ref_type = None
            for ref_type, base64_img in image_dict.items():
                if base64_img == matched_image:
                    matched_ref_type = ref_type
                    break
            print(f"  ✓ '{container_type}' -> '{matched_ref_type}' ({count} записей)")
        else:
            print(f"  ✗ '{container_type}' -> НЕ НАЙДЕНО ({count} записей)")
    
    # Сохраняем обновленный файл
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet, index=False)
    
    print(f"\nДанные сохранены в лист '{data_sheet}' с новым столбцом 'container_image_base64'")
    
    return df

def verify_base64_images(df, column='container_image_base64'):
    """
    Проверяет корректность base64 изображений
    """
    print(f"\nПроверка base64 изображений в столбце '{column}':")
    
    valid_count = 0
    invalid_count = 0
    
    for index, row in df.iterrows():
        base64_str = row[column] if column in row else None
        if pd.notna(base64_str) and isinstance(base64_str, str):
            try:
                # Пытаемся декодировать base64
                img_data = base64.b64decode(base64_str)
                
                # Проверяем, что это действительно изображение
                img = PILImage.open(io.BytesIO(img_data))
                valid_count += 1
                
                if index < 5:  # Показываем детали для первых 5 изображений
                    print(f"  Строка {index + 1}: ✓ {img.size} пикселей, формат {img.format}")
                    
            except Exception as e:
                invalid_count += 1
                print(f"  Строка {index + 1}: ✗ Ошибка: {e}")
    
    print(f"\nИтого: {valid_count} корректных изображений, {invalid_count} с ошибками")

# Альтернативная функция для извлечения изображений
def extract_images_alternative_approach(file_path, sheet_name="Справочник пробирок"):
    """
    Альтернативный подход к извлечению изображений - 
    сохраняем Excel как zip и извлекаем изображения из media папки
    """
    import zipfile
    import tempfile
    import os
    import shutil
    
    # Создаем временную папку
    with tempfile.TemporaryDirectory() as temp_dir:
        # Копируем Excel файл как zip
        zip_path = os.path.join(temp_dir, "excel_file.zip")
        shutil.copy2(file_path, zip_path)
        
        try:
            # Извлекаем zip архив
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Ищем папку с изображениями
            media_path = os.path.join(temp_dir, "xl", "media")
            if not os.path.exists(media_path):
                print("Папка с изображениями не найдена")
                return {}
            
            # Загружаем типы контейнеров из Excel
            df_ref = pd.read_excel(file_path, sheet_name=sheet_name)
            container_types = df_ref['Тип контейнера для хранения и транспортировки'].dropna().tolist()
            
            # Получаем все файлы изображений
            image_files = [f for f in os.listdir(media_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
            image_files.sort()  # Сортируем по имени файла
            
            print(f"Найдено файлов изображений: {len(image_files)}")
            print(f"Типов контейнеров: {len(container_types)}")
            
            image_dict = {}
            
            # Сопоставляем изображения с типами контейнеров по порядку
            for i, (image_file, container_type) in enumerate(zip(image_files, container_types)):
                try:
                    image_path = os.path.join(media_path, image_file)
                    
                    # Читаем изображение
                    with open(image_path, 'rb') as f:
                        img_data = f.read()
                    
                    # Конвертируем в PIL Image
                    pil_image = PILImage.open(io.BytesIO(img_data))
                    
                    # Изменяем размер для оптимизации
                    max_size = (300, 300)
                    pil_image.thumbnail(max_size, PILImage.Resampling.LANCZOS)
                    
                    # Конвертируем в base64
                    buffered = io.BytesIO()
                    pil_image.save(buffered, format="PNG")
                    img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                    
                    image_dict[container_type] = img_base64
                    print(f"  Сопоставлено: {container_type} -> {image_file} ({len(img_base64)} символов)")
                    
                except Exception as e:
                    print(f"  Ошибка при обработке {image_file}: {e}")
                    continue
            
            return image_dict
            
        except Exception as e:
            print(f"Ошибка при извлечении из zip: {e}")
            return {}

# Дополнительная функция для анализа совпадений
def analyze_matching_possibilities(file_path, image_dict, data_sheet="data"):
    """
    Анализирует возможные совпадения между названиями из data и справочника
    """
    # Читаем данные из листа 'data'
    df = pd.read_excel(file_path, sheet_name=data_sheet)
    
    # Получаем уникальные типы контейнеров из data
    data_container_types = df['container_type'].dropna().unique()
    
    # Получаем типы из справочника
    ref_container_types = list(image_dict.keys())
    
    print("=== Анализ возможных совпадений ===")
    print(f"Типы контейнеров в data: {len(data_container_types)}")
    print(f"Типы контейнеров в справочнике: {len(ref_container_types)}")
    
    print("\nТипы контейнеров в data:")
    for i, container_type in enumerate(data_container_types, 1):
        print(f"  {i}. {container_type}")
    
    print("\nТипы контейнеров в справочнике:")
    for i, container_type in enumerate(ref_container_types, 1):
        print(f"  {i}. {container_type}")
    
    print("\n=== Возможные совпадения ===")
    
    for data_type in data_container_types:
        data_type_clean = str(data_type).strip().lower()
        found_matches = []
        
        # Точное совпадение
        for ref_type in ref_container_types:
            if data_type_clean == ref_type.lower().strip():
                found_matches.append(('точное', ref_type))
        
        # Вхождение data в ref
        for ref_type in ref_container_types:
            ref_type_clean = ref_type.lower().strip()
            if data_type_clean in ref_type_clean:
                found_matches.append(('data в ref', ref_type))
        
        # Вхождение ref в data
        for ref_type in ref_container_types:
            ref_type_clean = ref_type.lower().strip()
            if ref_type_clean in data_type_clean:
                found_matches.append(('ref в data', ref_type))
        
        if found_matches:
            print(f"\n'{data_type}':")
            for match_type, ref_type in found_matches:
                print(f"  -> {match_type}: '{ref_type}'")
        else:
            print(f"\n'{data_type}': НЕ НАЙДЕНО совпадений")
    
    return data_container_types, ref_container_types

def main():
    project_dir = Path().resolve() # Path(__file__).resolve()
    # project_dir = project_dir.parent.parent
    raw_data_path = project_dir / 'data' / 'raw'
    processed_data_path = project_dir / 'data' / 'processed'
    file_path = raw_data_path / 'ВПР_Полный_перечень_тестов_МДЖ_23_06_ОКОНЧАТЕЛЬНЫЙ.xlsx'
    
    try:
        print("=== Извлечение изображений из справочника пробирок ===")
        
        # Сначала пробуем основной метод
        image_dict = extract_images_from_reference_sheet(file_path)
        
        # Если не получилось, пробуем альтернативный метод
        if not image_dict:
            print("\n=== Пробуем альтернативный метод ===")
            image_dict = extract_images_alternative_approach(file_path)
        
        if not image_dict:
            print("Не удалось извлечь изображения обоими способами.")
            return
        
        print(f"\n=== Найдено изображений: {len(image_dict)} ===")
        for container_type, base64_str in image_dict.items():
            print(f"  {container_type}: {len(base64_str)} символов")
        
        # Анализируем возможные совпадения
        print(f"\n=== Анализ совпадений ===")
        analyze_matching_possibilities(file_path, image_dict)
        
        # Добавляем столбец с base64 в лист "data"
        print(f"\n=== Добавление столбца в лист 'data' ===")
        df = add_base64_to_data_sheet(file_path, image_dict)
        
        # Проверяем результат
        verify_base64_images(df)
        
        print("\n=== Операция завершена успешно! ===")
        
    except Exception as e:
        print(f"Произошла ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

# Дополнительная функция для отладки - просмотр структуры листа
def debug_sheet_structure(file_path, sheet_name="Справочник пробирок"):
    """
    Отображает структуру листа для отладки
    """
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    print(f"=== Структура листа '{sheet_name}' ===")
    print(f"Размер листа: {worksheet.max_row} строк x {worksheet.max_column} столбцов")
    print(f"Количество изображений: {len(worksheet._images)}")
    
    # Показываем первые 10 строк
    print("\nПервые 10 строк:")
    for row in range(1, min(11, worksheet.max_row + 1)):
        values = []
        for col in range(1, min(5, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            values.append(str(cell_value)[:30] if cell_value else "")
        print(f"  Строка {row}: {' | '.join(values)}")
    
    # Информация об изображениях
    print(f"\nИнформация об изображениях:")
    for i, img in enumerate(worksheet._images):
        try:
            if hasattr(img.anchor, '_from'):
                row_idx = img.anchor._from.row + 1
                col_idx = img.anchor._from.col + 1
            else:
                row_idx = "неизвестно"
                col_idx = "неизвестно"
            print(f"  Изображение {i+1}: строка {row_idx}, столбец {col_idx}")
        except Exception as e:
            print(f"  Изображение {i+1}: ошибка определения позиции - {e}")
    
    workbook.close()

# Раскомментируйте для отладки:
# debug_sheet_structure(raw_data_path / 'ВПР_Полный_перечень_тестов_МДЖ_23_06_ОКОНЧАТЕЛЬНЫЙ.xlsx')